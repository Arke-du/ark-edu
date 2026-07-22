"""Importador das planilhas BNCC do Ensino Fundamental e Ensino Médio.

Compatível com planilhas no formato:

Ensino Fundamental:
    Disciplina | Ano | Unidade Temática | Objeto do Conhecimento | Habilidade

Ensino Médio:
    Disciplina | Ano | Habilidade

A célula "Habilidade" pode conter:
    (EF15AR01) Descrição...
ou:
    EF15AR01 - Descrição...

Uso:
    python importar_bncc.py __pycache__/BNCC_EF.xlsx --limpar
    python importar_bncc.py __pycache__/BNCC_EM.xlsx
"""

import argparse
import csv
import os
import re
import sqlite3
import unicodedata
from pathlib import Path

from bncc_catalogo import garantir_tabela_bncc


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.environ.get("DATABASE_PATH", BASE_DIR / "plataforma.db"))


def norm(valor):
    texto = str(valor or "").strip().casefold()
    return "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caractere) != "Mn"
    )


def chave_coluna(nome):
    return re.sub(r"[^a-z0-9]+", "_", norm(nome)).strip("_")


ALIASES = {
    "etapa_ensino": {
        "etapa",
        "etapa_de_ensino",
        "etapa_ensino",
    },
    "ano_serie": {
        "ano",
        "anos",
        "ano_serie",
        "ano_ou_bloco",
        "faixa_etaria",
    },
    "area_conhecimento": {
        "area",
        "area_de_conhecimento",
        "area_conhecimento",
    },
    "componente": {
        "componente",
        "componente_curricular",
        "disciplina",
    },
    "unidade_tematica": {
        "unidade_tematica",
        "pratica_de_linguagem",
        "campo_de_atuacao",
        "campo_de_atuacao_social",
    },
    "objeto_conhecimento": {
        "objeto_do_conhecimento",
        "objeto_de_conhecimento",
        "objeto_conhecimento",
        "objetos_do_conhecimento",
        "objetos_de_conhecimento",
    },
    "codigo": {
        "codigo",
        "codigo_da_habilidade",
        "codigo_habilidade",
    },
    "descricao": {
        "habilidade",
        "descricao_da_habilidade",
        "descricao",
        "texto_da_habilidade",
    },
}


def mapear_colunas(cabecalho):
    normalizadas = {
        chave_coluna(coluna): coluna
        for coluna in cabecalho
        if coluna is not None and str(coluna).strip()
    }

    mapa = {}
    for destino, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalizadas:
                mapa[destino] = normalizadas[alias]
                break

    # Nas planilhas fornecidas, etapa e código não aparecem como colunas
    # separadas. A etapa será inferida pelo código e o código será extraído
    # do começo do texto da habilidade.
    obrigatorias = {"componente", "descricao"}
    faltando = obrigatorias - mapa.keys()

    if faltando:
        encontradas = ", ".join(str(c) for c in cabecalho)
        raise ValueError(
            "Colunas obrigatórias não encontradas: "
            + ", ".join(sorted(faltando))
            + f". Cabeçalho encontrado: {encontradas}"
        )

    return mapa


def linha_parece_cabecalho(valores):
    chaves = {chave_coluna(v) for v in valores if v is not None}
    return (
        ("disciplina" in chaves or "componente" in chaves)
        and ("habilidade" in chaves or "descricao" in chaves)
    )


def ler_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(4096)
        arquivo.seek(0)

        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"

        linhas = list(csv.reader(arquivo, dialect=dialect))

    indice_cabecalho = None
    for indice, linha in enumerate(linhas[:30]):
        if linha_parece_cabecalho(linha):
            indice_cabecalho = indice
            break

    if indice_cabecalho is None:
        raise ValueError("Não foi possível localizar o cabeçalho no arquivo CSV.")

    cabecalho = [str(v or "").strip() for v in linhas[indice_cabecalho]]

    for valores in linhas[indice_cabecalho + 1 :]:
        if not any(str(v or "").strip() for v in valores):
            continue
        yield dict(zip(cabecalho, valores))


def ler_xlsx(path):
    """Lê planilhas e recompõe valores de células mescladas.

    Nas planilhas da BNCC, Disciplina, Ano, Unidade Temática e Objeto do
    Conhecimento podem estar em células mescladas. O openpyxl devolve o valor
    apenas na primeira linha da mesclagem e None nas linhas seguintes. Por isso,
    esta função mantém o último valor estrutural válido e o reaplica às linhas
    seguintes, respeitando a hierarquia dos campos.
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as erro:
        raise RuntimeError(
            "A biblioteca openpyxl não está instalada. Execute: "
            "python -m pip install openpyxl"
        ) from erro

    wb = load_workbook(path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        linhas = ws.iter_rows(values_only=True)
        cabecalho = None

        for numero_linha, valores in enumerate(linhas, start=1):
            if numero_linha > 30:
                break

            if linha_parece_cabecalho(valores):
                cabecalho = [str(v or "").strip() for v in valores]
                break

        if not cabecalho:
            print(f'Aviso: cabeçalho não localizado na aba "{ws.title}".')
            continue

        nomes_normalizados = {
            chave_coluna(nome): nome
            for nome in cabecalho
            if str(nome or "").strip()
        }

        coluna_disciplina = nomes_normalizados.get("disciplina") or nomes_normalizados.get("componente")
        coluna_ano = nomes_normalizados.get("ano") or nomes_normalizados.get("ano_serie")
        coluna_unidade = nomes_normalizados.get("unidade_tematica")
        coluna_objeto = (
            nomes_normalizados.get("objeto_do_conhecimento")
            or nomes_normalizados.get("objeto_conhecimento")
        )

        anteriores = {
            "disciplina": "",
            "ano": "",
            "unidade": "",
            "objeto": "",
        }

        for valores in linhas:
            if not any(str(v or "").strip() for v in valores):
                continue

            registro = {
                cabecalho[indice]: valores[indice] if indice < len(valores) else None
                for indice in range(len(cabecalho))
            }

            disciplina_atual = str(registro.get(coluna_disciplina, "") or "").strip() if coluna_disciplina else ""
            ano_atual = str(registro.get(coluna_ano, "") or "").strip() if coluna_ano else ""
            unidade_atual = str(registro.get(coluna_unidade, "") or "").strip() if coluna_unidade else ""
            objeto_atual = str(registro.get(coluna_objeto, "") or "").strip() if coluna_objeto else ""

            # Ao mudar um nível superior, evita reaproveitar indevidamente
            # valores pertencentes ao grupo anterior.
            if disciplina_atual and disciplina_atual != anteriores["disciplina"]:
                anteriores["ano"] = ""
                anteriores["unidade"] = ""
                anteriores["objeto"] = ""

            if ano_atual and ano_atual != anteriores["ano"]:
                anteriores["unidade"] = ""
                anteriores["objeto"] = ""

            if unidade_atual and unidade_atual != anteriores["unidade"]:
                anteriores["objeto"] = ""

            if disciplina_atual:
                anteriores["disciplina"] = disciplina_atual
            if ano_atual:
                anteriores["ano"] = ano_atual
            if unidade_atual:
                anteriores["unidade"] = unidade_atual
            if objeto_atual:
                anteriores["objeto"] = objeto_atual

            if coluna_disciplina and not disciplina_atual:
                registro[coluna_disciplina] = anteriores["disciplina"]
            if coluna_ano and not ano_atual:
                registro[coluna_ano] = anteriores["ano"]
            if coluna_unidade and not unidade_atual:
                registro[coluna_unidade] = anteriores["unidade"]
            if coluna_objeto and not objeto_atual:
                registro[coluna_objeto] = anteriores["objeto"]

            yield registro


PADRAO_HABILIDADE = re.compile(
    r"^\s*\(?\s*((?:EF|EM)[A-Z0-9]+)\s*\)?\s*[-–—:]?\s*(.+?)\s*$",
    flags=re.IGNORECASE | re.DOTALL,
)


def separar_codigo_descricao(valor_codigo, valor_habilidade):
    codigo = str(valor_codigo or "").strip().upper().replace(" ", "")
    texto = str(valor_habilidade or "").strip()

    if codigo:
        codigo = codigo.strip("()")
        descricao = texto
        descricao = re.sub(
            rf"^\s*\(?\s*{re.escape(codigo)}\s*\)?\s*[-–—:]?\s*",
            "",
            descricao,
            flags=re.IGNORECASE,
        ).strip()
        return codigo, descricao

    correspondencia = PADRAO_HABILIDADE.match(texto)
    if not correspondencia:
        return "", texto

    return (
        correspondencia.group(1).upper().replace(" ", ""),
        correspondencia.group(2).strip(),
    )


def normalizar_etapa(valor, codigo):
    etapa_informada = norm(valor)

    if codigo.startswith("EM") or "medio" in etapa_informada:
        return "Ensino Médio"

    if codigo.startswith("EF"):
        # EF15, EF12 e EF01-EF05 correspondem aos anos iniciais.
        faixa = codigo[2:4]

        if faixa in {"01", "02", "03", "04", "05", "12", "15"}:
            return "Ensino Fundamental - Anos Iniciais"

        if faixa in {"06", "07", "08", "09", "67", "69", "89"}:
            return "Ensino Fundamental - Anos Finais"

        # Alguns códigos valem para toda a etapa ou usam blocos atípicos.
        return "Ensino Fundamental"

    return str(valor or "").strip()


AREAS_EM = {
    "LGG": "Linguagens e suas Tecnologias",
    "LP": "Linguagens e suas Tecnologias",
    "MAT": "Matemática e suas Tecnologias",
    "CNT": "Ciências da Natureza e suas Tecnologias",
    "CHS": "Ciências Humanas e Sociais Aplicadas",
}


def prefixo_area_em(codigo):
    correspondencia = re.match(r"^EM\d{2}([A-Z]{2,3})", codigo)
    return correspondencia.group(1) if correspondencia else ""


def corrigir_componente_e_area(codigo, componente_planilha):
    componente_planilha = str(componente_planilha or "").strip()

    if not codigo.startswith("EM"):
        return componente_planilha, ""

    prefixo = prefixo_area_em(codigo)
    area = AREAS_EM.get(prefixo, componente_planilha)

    # No Ensino Médio, a maioria das habilidades pertence a uma área.
    # Língua Portuguesa possui códigos próprios EM..LP.
    if prefixo == "LP":
        componente = "Língua Portuguesa"
    else:
        componente = area

    return componente, area


def valor_campo(row, mapa, campo):
    nome_coluna = mapa.get(campo)
    if not nome_coluna:
        return ""
    return str(row.get(nome_coluna, "") or "").strip()


def importar(path, limpar=False):
    caminho = Path(path)

    if not caminho.exists():
        raise FileNotFoundError(
            f'Arquivo não encontrado: "{caminho}". '
            "Confira o nome e o caminho usando: find . -iname '*.xlsx'"
        )

    garantir_tabela_bncc(str(DB_PATH))

    if caminho.suffix.lower() in {".xlsx", ".xlsm"}:
        leitor = ler_xlsx(caminho)
    else:
        leitor = ler_csv(caminho)

    leitor = iter(leitor)

    try:
        primeira = next(leitor)
    except StopIteration as erro:
        raise ValueError(
            "A planilha está vazia ou não possui um cabeçalho reconhecível."
        ) from erro

    mapa = mapear_colunas(primeira.keys())

    def registros():
        yield primeira
        yield from leitor

    banco = sqlite3.connect(DB_PATH)

    try:
        if limpar:
            banco.execute("DELETE FROM bncc_habilidades")

        total_lidos = 0
        total_importados = 0
        total_ignorados = 0

        for row in registros():
            total_lidos += 1

            habilidade_bruta = valor_campo(row, mapa, "descricao")
            codigo_bruto = valor_campo(row, mapa, "codigo")

            codigo, descricao = separar_codigo_descricao(
                codigo_bruto,
                habilidade_bruta,
            )

            if not codigo or not descricao or not re.match(r"^(EF|EM)", codigo):
                total_ignorados += 1
                continue

            componente_planilha = valor_campo(row, mapa, "componente")
            componente, area_inferida = corrigir_componente_e_area(
                codigo,
                componente_planilha,
            )

            area_planilha = valor_campo(row, mapa, "area_conhecimento")
            area = area_planilha or area_inferida

            etapa = normalizar_etapa(
                valor_campo(row, mapa, "etapa_ensino"),
                codigo,
            )

            ano_serie = valor_campo(row, mapa, "ano_serie")
            unidade_tematica = valor_campo(row, mapa, "unidade_tematica")
            objeto_conhecimento = valor_campo(row, mapa, "objeto_conhecimento")

            banco.execute(
                """
                INSERT INTO bncc_habilidades (
                    etapa_ensino,
                    ano_serie,
                    area_conhecimento,
                    componente,
                    unidade_tematica,
                    objeto_conhecimento,
                    codigo,
                    descricao
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(
                    codigo,
                    componente,
                    ano_serie,
                    unidade_tematica,
                    objeto_conhecimento
                )
                DO UPDATE SET
                    descricao = excluded.descricao,
                    etapa_ensino = excluded.etapa_ensino,
                    area_conhecimento = excluded.area_conhecimento,
                    ativo = 1
                """,
                (
                    etapa,
                    ano_serie,
                    area,
                    componente,
                    unidade_tematica,
                    objeto_conhecimento,
                    codigo,
                    descricao,
                ),
            )

            total_importados += 1

        banco.commit()

        print()
        print("Importação concluída.")
        print(f"Arquivo: {caminho.resolve()}")
        print(f"Banco: {DB_PATH}")
        print(f"Linhas lidas: {total_lidos}")
        print(f"Registros importados/atualizados: {total_importados}")
        print(f"Linhas ignoradas: {total_ignorados}")

    except Exception:
        banco.rollback()
        raise
    finally:
        banco.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Importa planilhas da BNCC para o banco SQLite da ARK EDUS."
    )
    parser.add_argument("arquivo", help="Caminho do arquivo .xlsx, .xlsm ou .csv.")
    parser.add_argument(
        "--limpar",
        action="store_true",
        help="Apaga todo o catálogo BNCC anterior antes desta importação.",
    )

    argumentos = parser.parse_args()
    importar(argumentos.arquivo, argumentos.limpar)